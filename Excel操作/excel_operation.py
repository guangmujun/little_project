# -*- coding: utf-8 -*-
"""
Created on 2019/4/17 19:16

@author: WangYuhang

@function:
        纸质问卷信息录入电子表格中，平均1秒每题
        将下列q1q2q3列表的内容修改为对应问题的选项
        录入时只需输入题目选项的序号回车即可
"""

# 读写xlsx，安装openpyxl库
from openpyxl import load_workbook


def ope_excel(path, sheet_name, value, num):
    """
    向xlsx中写入数据后保存

    Parameters
    ----------
    path :
        xlsx文件的路径
    sheet_name :
        xlsx中表的名称
    value :
        输入的数据列表
    num :
        写入到表中的第几行

    Returns
    -------
        保存表格信息

    """
    wb = load_workbook(path)
    sheet = wb.get_sheet_by_name(sheet_name)

    for j in range(0, len(value)):
        sheet.cell(row=num + 1, column=j + 2, value=str(value[j]))

    wb.save(path)


q1 = [
    'A鼓楼',
    'B玄武',
    'C秦淮',
    'D河西',
    'E雨花',
    'F栖霞',
    'G板桥',
    'H江宁',
    'I浦口',
    'J高新大厂',
    'K六合',
    'L八卦洲',
    'M东部城市',
    'N南部城市',
    'O西部城市',
    'M北部城市']

q2 = ['步行', '私人自行车', '公共自行车', '电动自行车/摩托车', '公交', '地铁', '出租车/网约车', '私家车', '大巴']

q3 = ['0人', '1人', '2人', '3人', '4人', '5人', '6人', '7人', '7人以上']

path = r'.\data\调查问卷数据输入.xlsx'
sheet_name = '脑科医院门诊'

for i in range(0, 300):
    a1 = int(input("第 1 题答案："))
    a2 = int(input("第 2 题答案："))
    a3 = int(input("第 3 题答案："))
    value = [q1[a1], q2[a2], q3[a3]]
    num = i + 1

    ope_excel(path, sheet_name, value, num)
    print('--问卷 %s 完成--' % (num))
